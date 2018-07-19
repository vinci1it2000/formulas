#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Copyright 2016-2018 European Commission (JRC);
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl
import os
import unittest
import os.path as osp
import schedula as sh
from formulas.excel import ExcelModel, BOOK
from formulas.functions import is_number

EXTRAS = os.environ.get('EXTRAS', 'all')

mydir = osp.join(osp.dirname(__file__), 'test_files')
_filename = 'test.xlsx'
_filename_compile = 'excel.xlsx'
_link_filename = 'test_link.xlsx'


def _book2dict(book):
    res = {}
    for ws in book.worksheets:
        s = res[ws.title] = {}
        for k, cell in ws._cells.items():
            value = cell.value
            if value is not None:
                s[cell.coordinate] = value
    return res


@unittest.skipIf(EXTRAS not in ('all', 'excel'), 'Not for extra %s.' % EXTRAS)
class TestExcelModel(unittest.TestCase):
    def setUp(self):
        import openpyxl
        self.filename = osp.join(mydir, _filename)
        self.filename_compile = osp.join(mydir, _filename_compile)

        self.link_filename = osp.join(mydir, _link_filename)
        self.results = {
            _filename.upper(): _book2dict(
                openpyxl.load_workbook(self.filename, data_only=True)
            ),
            _link_filename.upper(): _book2dict(
                openpyxl.load_workbook(self.link_filename, data_only=True)
            ),
        }
        self.results_compile = _book2dict(
            openpyxl.load_workbook(self.filename_compile, data_only=True)
        )['DATA']

        self.maxDiff = None

    def _compare(self, books, results):
        it = sorted(sh.stack_nested_keys(results, depth=3))
        for k, res in it:
            value = sh.get_nested_dicts(books, *k)
            msg = '[{}]{}!{}'.format(*k)
            if is_number(value) and is_number(res):
                self.assertAlmostEqual(float(res), float(value), msg=msg)
            else:
                self.assertEqual(res, value, msg=msg)
        return len(it)

    def test_excel_model(self):
        xl_model = ExcelModel()
        xl_model.loads(self.filename)
        xl_model.add_book(self.link_filename)
        xl_model.finish()
        xl_model.calculate()
        books = xl_model.books
        books = {k: _book2dict(v[BOOK])
                 for k, v in xl_model.write(books).items()}

        n_test = self._compare(books, self.results)

        books = {k: _book2dict(v[BOOK]) for k, v in xl_model.write().items()}
        res = {}
        for k, v in sh.stack_nested_keys(self.results, depth=2):
            sh.get_nested_dicts(res, *map(str.upper, k), default=lambda: v)

        n_test += self._compare(books, res)
        print('[info] test_excel_model: Ran %d tests.' % n_test)

    def test_excel_model_compile(self):
        xl_model = ExcelModel()
        xl_model.loads(self.filename_compile)
        xl_model.finish()
        inputs = ["A%d" % i for i in range(2, 5)]
        outputs = ["C%d" % i for i in range(2, 5)]
        func = xl_model.compile(
            ["'[EXCEL.XLSX]DATA'!%s" % i for i in inputs],
            ["'[EXCEL.XLSX]DATA'!%s" % i for i in outputs]
        )
        i = sh.selector(inputs, self.results_compile, output_type='list')
        res = sh.selector(outputs, self.results_compile, output_type='list')
        self.assertEqual([x.value[0, 0] for x in func(*i)], res)
