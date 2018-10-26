#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Copyright 2016-2018 European Commission (JRC);
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl

"""
It provides Excel model class.
"""

import os.path as osp
import schedula as sh
import dill

from schedula.utils.io import save_dispatcher, load_dispatcher

from formulas.errors import FormulaError
from .ranges import Ranges
from .cell import Cell, RangesAssembler
from .tokens.operand import range2parts, XlError
from .functions import flatten

BOOK = sh.Token('Book')
SHEETS = sh.Token('Sheets')
CIRCULAR = sh.Token('CIRCULAR')
ERR_CIRCULAR = XlError('0')


def _get_name(name, names):
    if name not in names:
        name = name.upper()
        for n in names:
            if n.upper() == name:
                return n
    return name


class ExcelModel(object):
    compile_class = sh.DispatchPipe

    def __init__(self):
        self.dsp = sh.Dispatcher()
        self.calculate = self.dsp.dispatch
        self.cells = {}
        self.books = {}

    @staticmethod
    def _yield_refs(book, context=None):
        for n in book.defined_names.definedName:
            if n.value == '#REF!':
                continue
            ref, i = n.name.upper(), n.localSheetId
            rng = Ranges().push(n.value, context=context).ranges[0]['name']
            sheet_names = book.sheetnames
            if i is not None:
                sheet_names = sheet_names[i:i + 1]
            for sn in sheet_names:
                name = range2parts(
                    None, **sh.combine_dicts(context, {'sheet': sn, 'ref': ref})
                )
                yield name['name'], rng

    def loads(self, *file_names):
        for filename in file_names:
            self.load(filename)
        return self

    def load(self, filename):
        book, context = self.add_book(filename)
        self.pushes(*book.worksheets, context=context)
        return self


    def _load_bin(self, fn):
        with open(fn, 'rb') as f:
            return dill.load(f)

    def _save_bin(self, obj, fn):
        with open(fn, 'wb') as f:
            dill.dump(obj, f)

    def save_bin(self, fn):
        self._save_bin(self, fn + ".fxl")

    def _update_constant_value(self, attr, value):
        setattr(sh.utils.cst, attr, value)
        setattr(sh.utils.sol, attr, value)
        setattr(sh.dispatcher, attr, value)
        setattr(sh.utils.alg, attr, value)
        setattr(sh.utils, attr, value)
        setattr(sh.utils.base, attr, value)
        setattr(sh, attr, value)

    def _update_constants(self, constant_token):
        k = constant_token
        if(str(k) == "start"):
            self._update_constant_value("START", k)
        if(str(k) == "empty"):
            self._update_constant_value("EMPTY", k)
        if(str(k) == "none"):
            self._update_constant_value("NONE", k)
        if(str(k) == "sink"):
            self._update_constant_value("SINK", k)
        if(str(k) == "end"):
            self._update_constant_value("END", k)
        if(str(k) == "self"):
            self._update_constant_value("SELF", k)
        if(str(k) == "plot"):
            self._update_constant_value("PLOT", k)

    def _populate_self(self, xl_model):
        self.dsp = xl_model.dsp
        self.cells = xl_model.cells
        self.books = xl_model.books
        self.calculate = self.dsp.dispatch
        for k, v in self.dsp.nodes.items():
            self._update_constants(k)
            self._update_constants(v)
        for k, v in self.cells.items():
            if(str(v.range._value)=="none"):
                self._update_constants(v.range._value)
                break

    def load_bin_file(self, f):
        """Load already open filehandle"""
        xl_model = dill.load(f)
        self._populate_self(xl_model)

    def load_bin(self, fn):
        """Load file by name"""
        xl_model = self._load_bin(fn + ".fxl")
        self._populate_self(xl_model)

    def pushes(self, *worksheets, context=None):
        for ws in worksheets:
            self.push(ws, context=context)
        return self

    def push(self, worksheet, context):
        worksheet, context = self.add_sheet(worksheet, context)

        get_in = sh.get_nested_dicts
        references = get_in(self.books, context['excel'], 'references')
        d = get_in(self.books, context['excel'], SHEETS, context['sheet'])
        formula_references = d['formula_references']
        formula_ranges = d['formula_ranges']

        for row in worksheet.iter_rows():
            for c in row:
                self.add_cell(
                    c, context, references=references,
                    formula_references=formula_references,
                    formula_ranges=formula_ranges
                )
        return self

    def add_book(self, book, context=None, data_only=False):
        context = context or {}
        are_in = sh.are_in_nested_dicts
        get_in = sh.get_nested_dicts

        if 'excel' in context:
            context = context.copy()
            context['excel'] = context['excel'].upper()

        if are_in(self.books, context.get('excel'), BOOK):
            book = get_in(self.books, context['excel'], BOOK)
        else:
            if isinstance(book, str):
                context.update({'excel': osp.basename(book).upper(),
                                'directory': osp.dirname(osp.abspath(book))})
                if not are_in(self.books, context['excel'], BOOK):
                    from openpyxl import load_workbook
                    book = load_workbook(book, data_only=data_only)
            book = get_in(
                self.books, context['excel'], BOOK, default=lambda: book
            )

        if not are_in(self.books, context['excel'], 'references'):
            get_in(
                self.books, context['excel'], 'references',
                default=lambda: dict(self._yield_refs(book, context=context))
            )

        if not are_in(self.books, context['excel'], 'external_links'):
            external_links = {str(l.file_link.idx_base + 1): l.file_link.Target
                              for l in book._external_links}
            get_in(
                self.books, context['excel'], 'external_links',
                default=lambda: external_links
            )

        return book, context

    def add_sheet(self, worksheet, context):
        get_in = sh.get_nested_dicts
        if isinstance(worksheet, str):
            book = get_in(self.books, context['excel'], BOOK)
            worksheet = book[_get_name(worksheet, book.sheetnames)]

        context = sh.combine_dicts(
            context, base={'sheet': worksheet.title.upper()}
        )

        d = get_in(self.books, context['excel'], SHEETS, context['sheet'])
        if 'formula_references' not in d:
            d['formula_references'] = formula_references = {
                k: v['ref'] for k, v in worksheet.formula_attributes.items()
                if v.get('t') == 'array' and 'ref' in v
            }
        else:
            formula_references = d['formula_references']

        if 'formula_ranges' not in d:
            d['formula_ranges'] = {
                Ranges().push(ref, context=context)
                for ref in formula_references.values()
            }
        return worksheet, context

    def add_cell(self, cell, context, references=None, formula_references=None,
                 formula_ranges=None, external_links=None):
        get_in = sh.get_nested_dicts
        if formula_references is None:
            formula_references = get_in(
                self.books, context['excel'], SHEETS, context['sheet'],
                'formula_references'
            )

        if formula_ranges is None:
            formula_ranges = get_in(
                self.books, context['excel'], SHEETS, context['sheet'],
                'formula_ranges'
            )

        if references is None:
            references = get_in(self.books, context['excel'], 'references')

        if external_links is None:
            external_links = get_in(
                self.books, context['excel'], 'external_links'
            )
        context = sh.combine_dicts(
            context, base={'external_links': external_links}
        )
        crd = cell.coordinate
        crd = formula_references.get(crd, crd)
        try:
            cell = Cell(crd, cell.value, context=context).compile()
        except FormulaError:
            # There was an error so set this value with the NA error using the excel function
            cell = Cell(crd, "=NA()", context=context).compile()
        if cell.output in self.cells:
            return
        if cell.value is not sh.EMPTY:
            if any(not (cell.range - rng).ranges for rng in formula_ranges):
                return
        cell.update_inputs(references=references)

        if cell.add(self.dsp, context=context):
            self.cells[cell.output] = cell
            return cell

    def complete(self):
        nodes = self.dsp.nodes
        stack = list(sorted(set(self.dsp.data_nodes) - set(self.cells)))
        while stack:
            n_id = stack.pop()
            if isinstance(n_id, sh.Token):
                continue

            rng = Ranges().push(n_id).ranges[0]
            book = osp.abspath(
                osp.join(nodes[n_id].get('directory', '.'), rng['excel'])
            )
            context = self.add_book(book)[1]
            worksheet, context = self.add_sheet(rng['sheet'], context)
            rng = '{c1}{r1}:{c2}{r2}'.format(**rng)
            for c in flatten(worksheet[rng], None):
                cell = self.add_cell(c, context)
                if cell:
                    stack.extend(cell.inputs or ())

    def finish(self, complete=True, circular=False):
        if complete:
            self.complete()
        for n_id in sorted(set(self.dsp.data_nodes) - set(self.cells)):
            if isinstance(n_id, sh.Token):
                continue
            ra = RangesAssembler(n_id)
            for k, c in sorted(self.cells.items()):
                ra.push(c)
                if not ra.missing.ranges:
                    break

            self.dsp.add_function(None, ra, ra.inputs or None, [ra.output])

        if circular:
            self.solve_circular()

        return self

    def write(self, books=None, solution=None):
        books = {} if books is None else books
        solution = self.dsp.solution if solution is None else solution
        are_in = sh.are_in_nested_dicts
        get_in = sh.get_nested_dicts
        for k, r in solution.items():
            if isinstance(k, sh.Token):
                continue
            rng = r.ranges[0]
            filename, sheet_name = _get_name(rng['excel'], books), rng['sheet']

            if not are_in(books, filename, BOOK):
                from openpyxl import Workbook
                book = get_in(books, filename, BOOK, default=Workbook)
                for ws in book.worksheets:
                    book.remove(ws)
            else:
                book = books[filename][BOOK]

            sheet_names = book.sheetnames
            sheet_name = _get_name(sheet_name, sheet_names)
            if sheet_name not in sheet_names:
                book.create_sheet(sheet_name)
            sheet = book[sheet_name]

            ref = '{c1}{r1}:{c2}{r2}'.format(**rng)
            for c, v in zip(flatten(sheet[ref], None), flatten(r.value, None)):
                if v is sh.EMPTY:
                    v = None
                c.value = v

        return books

    def compile(self, inputs, outputs):
        dsp = self.dsp.shrink_dsp(outputs=outputs)

        dsp.default_values = sh.selector(
            set(dsp.default_values) - set(inputs), dsp.default_values
        )

        res = dsp()

        dsp = dsp.get_sub_dsp_from_workflow(
            outputs, graph=dsp.dmap, reverse=True, blockers=res,
            wildcard=False
        )

        for k, v in sh.selector(dsp.data_nodes, res, allow_miss=True).items():
            try:
                dsp.set_default_value(k, v.value)
            except AttributeError:
                # Circular token does not have v.value (it is an XlError)
                pass
        func = self.compile_class(
            dsp=dsp,
            function_id=self.dsp.name,
            inputs=inputs,
            outputs=outputs
        )

        return func

    def solve_circular(self):
        import networkx as nx
        mod, dsp = {}, self.dsp
        f_nodes, d_nodes, dmap = dsp.function_nodes, dsp.data_nodes, dsp.dmap

        for cycle in sorted(map(set, nx.simple_cycles(dmap))):
            for k in sorted(cycle.intersection(f_nodes)):
                if _check_cycles(dmap, k, f_nodes, cycle, mod):
                    break
            else:
                dist = sh.inf(len(cycle) + 1, 0)
                for k in sorted(cycle.intersection(d_nodes)):
                    dsp.set_default_value(k, ERR_CIRCULAR, dist)

        if mod:  # Update dsp.
            dsp.add_data(CIRCULAR, ERR_CIRCULAR)

            for k, v in mod.items():
                d = f_nodes[k]
                d['inputs'] = [CIRCULAR if i in v else i for i in d['inputs']]
                dmap.remove_edges_from(((i, k) for i in v))
                dmap.add_edge(CIRCULAR, k)

        return self


def _check_cycles(dmap, node_id, nodes, cycle, mod=None):
    node, mod = nodes[node_id], {} if mod is None else mod
    _map = dict(zip(node['function'].inputs, node['inputs']))
    pred, res = dmap.predecessors, ()
    check = lambda j: isinstance(nodes[j]['function'], RangesAssembler)
    if not any(any(map(check, pred(k))) for k in _map.values() if k in cycle):
        cycle = [i for i, j in _map.items() if j in cycle]
        try:
            res = tuple(map(_map.get, node['function'].check_cycles(cycle)))
            res and sh.get_nested_dicts(mod, node_id, default=set).update(res)
        except AttributeError:
            pass
    return res
