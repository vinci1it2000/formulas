import regex
import ezodf
import os.path as osp
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula

_bracket_ref = regex.compile(r"\[([^\]]+)\]")
_of_prefix = regex.compile(r"^\s*of:(?==)")

_re_clean = regex.compile(
    r'"[^"]*"(*SKIP)(*F)|\.(?=(\$?[A-Z]+\$?\d+|#))|\$\$(?=[_A-Z])|COM\.MICROSOFT\.|LEGACY\.',
    regex.IGNORECASE
)
_re_replace = regex.compile(
    r'"[^"]*"(*SKIP)(*F)|(?<=[A-Z0-9\'])\.(?=(\$?[A-Z]+\$?\d+|#))',
    regex.IGNORECASE
)
_external_link = regex.compile(
    r"('file:[^']*')#(?:\$?([^!]+)!)?",
    regex.IGNORECASE,
)

_range_same_sheet = regex.compile(
    r"(?i)^(?P<sheet>(?:'[^']+'|\[[^\]]+\]|[^'!\[]+)+)!"
    r"(?P<left>\$?[A-Z]{1,3}\$?\d+):"
    r"(?P=sheet)!"
    r"(?P<right>\$?[A-Z]{1,3}\$?\d+)$",
    regex.IGNORECASE,
)


def _external_link_to_excel(string, links):
    def _sub_external_link(m):
        g1, g2 = m.groups()
        return f"'{links[g1]}{g2}'!"

    return _external_link.sub(_sub_external_link, string)


def _odf_range_to_excel(addr: str) -> str:
    return _range_same_sheet.sub(
        r"\g<sheet>!\g<left>:\g<right>",
        _re_clean.sub('', _re_replace.sub('!', addr))
    )


def replace_semicolon_outside_quotes(expr: str) -> str:
    """Replace ; with , only when outside double quotes."""
    result = []
    in_quotes = False
    in_array = False
    i = 0
    while i < len(expr):
        ch = expr[i]
        if ch == '"':
            result.append(ch)
            in_quotes = not in_quotes
        elif not in_array and (
                (ch == "{" and not in_quotes) or (ch == "}" and in_quotes)
        ):
            result.append(ch)
            in_array = not in_array
        elif ch in ";~" and not in_quotes:
            result.append(",")
        elif ch == "!" and expr[i - 1] in "])" and not in_quotes:
            result.append(" ")
        elif ch == "|" and not in_quotes and in_array:
            result.append(";")
        else:
            result.append(ch)
        i += 1
    return "".join(result)


def translate_odf_formula_to_excel(odf_formula: str, links: dict) -> str:
    """
    Convert an ODF formula like:
      'of:=SUM([.A1:.A10])'
    to an Excel formula like:
      '=SUM(Sheet!A1:A10)'
    """
    if odf_formula:
        # drop of:=
        expr = _of_prefix.sub('', odf_formula)

        # Replace argument separators ';' -> ',' (common in ODF locales)
        # This is heuristic; if your locale already uses ',', this is harmless.
        expr = replace_semicolon_outside_quotes(expr)

        # Replace each bracketed reference [ ... ] with an Excel-style ref
        def _sub(m):
            return _odf_range_to_excel(m.group(1))

        expr = _bracket_ref.sub(_sub, expr)
        expr = _re_replace.sub('!', expr)
        expr = _external_link_to_excel(_re_clean.sub('', expr), links)

        return expr


def ods_to_xlsx(ods_path: str, data_only=False, **kwargs):
    """
    Convert an ODS workbook to XLSX (openpyxl), preserving values, formulas,
    named ranges/expressions, and (optionally) database ranges as Excel Tables.
    """
    from ezodf.xmlns import CN
    from . import _decode_path, _encode_path
    doc = ezodf.opendoc(ods_path)
    wb = Workbook()
    # remove default first sheet; we'll add sheets according to ODS order
    default_ws = wb.active
    wb.remove(default_ws)
    basedir = osp.dirname(ods_path)
    links = {}
    for sheet in doc.sheets:
        src = sheet.xmlnode.find(CN('table:table-source'))
        if src is not None:
            fdir, fname = osp.split(osp.relpath(osp.realpath(osp.join(
                ods_path, _decode_path(src.get(CN('xlink:href')))
            )), basedir))
            links[sheet.name.split('#')[0]] = _encode_path(osp.join(
                fdir, f'[{fname}]'
            ))

    # --- build a map of sheet name -> openpyxl worksheet ---

    for sheet in doc.sheets:
        sn = sheet.name
        if sheet.xmlnode.find(CN('table:table-source')) is not None:
            continue
        ws = wb.create_sheet(title=sn)

        for irow, row in enumerate(sheet.rows(), 1):
            for icol, cell in enumerate(row, 1):
                if cell.formula and not data_only:
                    value = translate_odf_formula_to_excel(
                        cell.formula, links
                    )
                    rows = cell.xmlnode.get(
                        CN('table:number-matrix-rows-spanned')
                    )
                    cols = cell.xmlnode.get(
                        CN('table:number-matrix-columns-spanned')
                    )
                    if rows or cols:
                        end_row = irow + int(rows or 1) - 1
                        end_col = icol + int(cols or 1) - 1
                        value = ArrayFormula(
                            f"{get_column_letter(icol)}{irow}:{get_column_letter(end_col)}{end_row}",
                            value
                        )
                elif cell.display_form == 'Err:502':
                    value = '#N/A'
                elif cell.value_type == 'error':
                    value = cell.display_form
                else:
                    value = cell.value

                if value is not None:
                    ws.cell(row=irow, column=icol, value=value)
                    if cell.span != (1, 1):
                        ws.merge_cells(
                            start_row=irow,
                            start_column=icol,
                            end_row=irow + cell.span[0] - 1,
                            end_column=icol + cell.span[1] - 1
                        )

    for block in doc.body.findall(CN("table:named-expressions")):
        # Named ranges (<table:named-range> with cell-range-address)
        for nr in block.findall(CN("table:named-range")):
            name = nr.get_attr(CN("table:name"))
            addr = nr.get_attr(CN("table:cell-range-address"))
            if name and addr:
                excel_ref = _external_link_to_excel(
                    _odf_range_to_excel(addr), links
                )
                wb.defined_names[name] = DefinedName(
                    name=name, attr_text=excel_ref
                )

        # Named expressions (<table:named-expression table:expression="of:=...">)
        for ne in block.findall(CN("table:named-expression")):
            name = ne.get_attr(CN("table:name"))
            expr = ne.get_attr(CN("table:expression"))
            if name and expr:
                xl_expr = translate_odf_formula_to_excel(expr, links)
                wb.defined_names[name] = DefinedName(
                    name=name,
                    attr_text=xl_expr[1:] if xl_expr[0] == '=' else xl_expr
                )

    return wb
