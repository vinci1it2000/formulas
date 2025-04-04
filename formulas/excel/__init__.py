#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Copyright 2016-2025 European Commission (JRC);
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl

"""
It provides Excel model class.

Sub-Modules:

.. currentmodule:: formulas.excel

.. autosummary::
    :nosignatures:
    :toctree: excel/

    ~cycle
    ~xlreader
"""
import os
import logging
import functools
import numpy as np
import os.path as osp
import schedula as sh
from ..ranges import Ranges
from ..errors import InvalidRangeName
from ..cell import Cell, RangesAssembler, Ref, CellWrapper, InvRangesAssembler
from ..tokens.operand import XlError, _re_sheet_id, _re_build_id
from ..functions.text import HexValue

log = logging.getLogger(__name__)
BOOK = sh.Token('Book')
SHEETS = sh.Token('Sheets')
CIRCULAR = sh.Token('CIRCULAR')


class XlCircular(XlError):
    def __str__(self):
        return '0'


ERR_CIRCULAR = XlCircular('#CIRC!')


def _get_name(name, names):
    if name not in names:
        name = name.upper()
        for n in names:
            if n.upper() == name:
                return n
    return name


def _encode_path(path):
    return path.replace('\\', '/')


def _decode_path(path):
    return path.replace('/', osp.sep)


def _book2dict(book):
    res = {}
    for ws in book.worksheets:
        s = res[ws.title.upper()] = {}
        for k, cell in ws._cells.items():
            value = getattr(cell, 'value', None)
            if value is not None:
                s[cell.coordinate] = value
    return res


def _res2books(res):
    return {k.upper(): _book2dict(v[BOOK]) for k, v in res.items()}


def _file2books(*fpaths):
    from .xlreader import load_workbook
    d = osp.dirname(fpaths[0])
    return {osp.relpath(fp, d).upper().replace('\\', '/'): _book2dict(
        load_workbook(fp, data_only=True)
    ) for fp in fpaths}


class ExcelModel:
    compile_class = sh.DispatchPipe

    def __init__(self):
        self.dsp = sh.Dispatcher(name='ExcelModel')
        self.cells = {}
        self.books = {}
        self.basedir = None

    def __call__(self, *args, **kwargs):
        return self.calculate(*args, **kwargs)

    def calculate(self, *args, **kwargs):
        return self.dsp.dispatch(*args, **kwargs)

    def compare(self, *fpaths, solution=None, tolerance=.000001, **kwargs):
        if solution is None:
            solution = self.dsp.dispatch()
        from dictdiffer import diff
        target = _file2books(*fpaths)
        return list(diff(target, sh.selector(
            target, _res2books(self.write(solution=solution))
        ), tolerance=0, absolute_tolerance=tolerance, **kwargs))

    def __getstate__(self):
        return {'dsp': self.dsp, 'cells': {}, 'books': {}}

    def _update_refs(self, nodes, refs):
        if nodes:
            dsp = self.dsp.get_sub_dsp(nodes)
            dsp.raises = ''
            sol = dsp({k: sh.EMPTY for k in dsp.data_nodes if k not in refs})
            refs.update({
                k: v for k, v in sol.items()
                if k in refs and isinstance(v, Ranges)
            })

    def add_references(self, book, context=None):
        refs, nodes = {}, set()
        it = book.defined_names
        it = it.values() if isinstance(it, dict) else it.definedName
        for n in it:
            if n.hidden or n.localSheetId is not None:
                continue  # Accepts only global references.
            ref = Ref(n.name.upper(), '=%s' % n.value, context).compile(
                context=context
            )
            nodes.update(ref.add(self.dsp, context=context))
            refs[ref.output] = None
            self.cells[ref.output] = ref
        self._update_refs(nodes, refs)
        return refs

    def loads(self, *file_names):
        for filename in file_names:
            self.load(filename)
        return self

    def load(self, filename):
        book, context = self.add_book(filename)
        self.pushes(*book.worksheets, context=context)
        return self

    def from_ranges(self, *ranges):
        return self.complete(ranges)

    def pushes(self, *worksheets, context=None):
        for ws in worksheets:
            self.push(ws, context=context)
        return self

    def push(self, worksheet, context):
        worksheet, context = self.add_sheet(worksheet, context)
        references = self.references
        formula_references = self.formula_references(context)
        formula_ranges = self.formula_ranges(context)
        external_links = self.external_links(context)
        ctx = {'external_links': external_links}
        ctx.update(context)
        cells = []
        for row in worksheet.iter_rows():
            for c in row:
                if hasattr(c, 'value'):
                    cells.append(self.compile_cell(
                        c, ctx, references, formula_references
                    ))
        for cell in cells:
            # noinspection PyTypeChecker
            self.add_cell(sh.await_result(cell), ctx, formula_ranges)
        return self

    def add_book(self, book=None, context=None, data_only=False):
        ctx = (context or {}).copy()
        are_in, get_in = sh.are_in_nested_dicts, sh.get_nested_dicts
        if isinstance(book, str):
            book = _encode_path(book).split('/')
            ctx['directory'], ctx['filename'] = '/'.join(book[:-1]), book[-1]
        if self.basedir is None:
            directory = _encode_path(ctx.get('directory') or '.')
            self.basedir = osp.abspath(_decode_path(directory))
            ctx['directory'] = ''
        if ctx['directory']:
            ctx['directory'] = _encode_path(osp.relpath(
                osp.join(self.basedir, _decode_path(ctx['directory'])),
                self.basedir
            ))
        if ctx['directory'] == '.':
            ctx['directory'] = ''
        fpath = osp.join(_decode_path(ctx['directory']), ctx['filename'])
        ctx['excel'] = _encode_path(fpath).upper()
        data = get_in(self.books, ctx['excel'])
        book = data.get(BOOK)
        if not book:
            from .xlreader import load_workbook
            data[BOOK] = book = load_workbook(
                osp.join(self.basedir, fpath), data_only=data_only
            )

        if 'external_links' not in data:
            fdir = osp.join(self.basedir, _decode_path(ctx['directory']))
            data['external_links'] = {
                str(i + 1): osp.split(osp.relpath(osp.realpath(osp.join(
                    fdir, _decode_path(el.file_link.Target)
                )), self.basedir))
                for i, el in enumerate(book._external_links)
                if el.file_link.Target.endswith('.xlsx')
            }
            data['external_links'] = {
                k: (_encode_path(d), f)
                for k, (d, f) in data['external_links'].items()
            }

        if 'references' not in data:
            context = {'external_links': data['external_links']}
            context.update(ctx)
            data['references'] = self.add_references(book, context=context)

        return book, ctx

    def add_sheet(self, worksheet, context):
        get_in = sh.get_nested_dicts
        if isinstance(worksheet, str):
            book = get_in(self.books, context['excel'], BOOK)
            worksheet = book[_get_name(worksheet, book.sheetnames)]

        ctx = {'sheet': worksheet.title.upper()}
        ctx.update(context)

        d = get_in(self.books, ctx['excel'], SHEETS, ctx['sheet'])
        if 'formula_references' not in d:
            try:
                formula_references = {
                    k: v['ref'] for k, v in worksheet.formula_attributes.items()
                    if v.get('t') == 'array' and 'ref' in v
                }
            except AttributeError:  # openpyxl>=3.1
                formula_references = worksheet.array_formulae.copy()

            d['formula_references'] = formula_references
        else:
            formula_references = d['formula_references']

        if 'formula_ranges' not in d:
            d['formula_ranges'] = {
                Ranges().push(ref, context=ctx)
                for ref in formula_references.values()
            }
        return worksheet, ctx

    @property
    def references(self):
        return sh.combine_dicts(*(
            d.get('references', {}) for d in self.books.values()
        ))

    def formula_references(self, ctx):
        return sh.get_nested_dicts(
            self.books, ctx['excel'], SHEETS, ctx['sheet'], 'formula_references'
        )

    def formula_ranges(self, ctx):
        return sh.get_nested_dicts(
            self.books, ctx['excel'], SHEETS, ctx['sheet'], 'formula_ranges'
        )

    def external_links(self, ctx):
        return sh.get_nested_dicts(self.books, ctx['excel'], 'external_links')

    @staticmethod
    def _compile_cell(crd, val, context, check_formula, references):
        cell = Cell(crd, val, context=context, check_formula=check_formula)
        cell.compile(references=references, context=context)
        return cell

    def compile_cell(self, cell, context, references, formula_references):
        crd = cell.coordinate
        crd = formula_references.get(crd, crd)
        val = cell.value
        if cell.data_type == 'f':
            if not isinstance(val, str):
                val = val.text
            val = val[:2] == '==' and val[1:] or val
        elif cell.data_type == 'n' and isinstance(val, float):
            val = round(val, 15)

        check_formula = cell.data_type != 's'
        return self._compile_cell(crd, val, context, check_formula, references)

    def add_cell(self, cell, context, formula_ranges):
        if cell.output in self.cells:
            return
        if cell.value is not sh.EMPTY:
            if any(not (cell.range - rng).ranges for rng in formula_ranges):
                return

        if cell.add(self.dsp, context=context):
            self.cells[cell.output] = cell
            return cell

    def complete(self, stack=None):
        done = set(self.cells)
        if stack is None:
            stack = {k for k in self.dsp.data_nodes if k not in self.references}
            stack = stack.difference(done)
        stack = sorted(stack)
        sheet_limits = {}
        while stack:
            n_id = stack.pop()
            if isinstance(n_id, sh.Token) or n_id in done:
                continue
            done.add(n_id)
            if n_id in self.references:
                stack.extend(self.cells[n_id].inputs or ())
                continue
            try:
                rng = Ranges.get_range(n_id, raise_anchor=False)
            except InvalidRangeName:  # Missing Reference.
                log.warning('Missing Reference `{}`!'.format(n_id))
                Ref(n_id, '=#REF!').compile().add(self.dsp)
                continue
            book = _encode_path(osp.join(
                _decode_path(rng.get('directory', '')),
                _decode_path(rng.get('filename', rng.get('excel_id', '')))
            ))

            try:
                context = self.add_book(book)[1]
                wk, context = self.add_sheet(rng['sheet'], context)
            except Exception as ex:  # Missing excel file or sheet.
                log.warning('Error in loading `{}`:\n{}'.format(n_id, ex))
                Cell(n_id, '=#REF!').compile().add(self.dsp)
                self.books.pop(book, None)
                continue
            formula_references = self.formula_references(context)
            if rng.get('anchor'):
                ref = formula_references.get(f"{rng['c1']}{rng['r1']}")
                if ref:
                    ref = Ranges.get_range(ref, context)['name']
                    self.dsp.add_function(
                        function_id=f'={ref}',
                        function=sh.bypass,
                        inputs=[ref],
                        outputs=[n_id]
                    )
                    stack.append(ref)
                else:
                    Cell(n_id, '=#REF!').compile().add(self.dsp)
                continue
            references = self.references
            formula_ranges = self.formula_ranges(context)
            external_links = self.external_links(context)

            _name = '%s'
            if 'sheet_id' in rng:
                _name = f'{rng["sheet_id"]}!{_name}'
            if wk not in sheet_limits:
                sheet_limits[wk] = wk.max_row, wk.max_column
            max_row, max_column = sheet_limits[wk]
            it = wk.iter_rows(
                int(rng['r1']), min(int(rng['r2']), max_row),
                rng['n1'], min(rng['n2'], max_column)
            )
            ctx = {'external_links': external_links}
            ctx.update(context)
            cells = []
            for row in it:
                for c in row:
                    n = _name % c.coordinate
                    if n in self.cells:
                        continue
                    elif hasattr(c, 'value'):
                        cells.append(self.compile_cell(
                            c, ctx, references, formula_references
                        ))
            for cell in cells:
                # noinspection PyTypeChecker
                cell = self.add_cell(sh.await_result(cell), ctx, formula_ranges)
                if cell:
                    stack.extend(cell.inputs or ())
        return self

    def _assemble_ranges(self, cells, nodes=None, compact=1):
        get, dsp = sh.get_nested_dicts, self.dsp
        pred = dsp.dmap.pred
        if nodes is None:
            nodes = {k for k in dsp.data_nodes if k not in dsp.default_values}
        it = (
            k for k in nodes
            if not pred[k] and not isinstance(k, sh.Token)
        )
        ranges = []
        for n_id in it:
            try:
                ra = RangesAssembler(n_id, compact=compact)
            except ValueError:
                continue
            rng = ra.range.ranges[0]
            for out, idx in get(cells, 'range', rng['sheet_id'], default=list):
                if not ra.push(idx, out):
                    break
            ra.push(get(cells, 'cell', rng['sheet_id']))
            ranges.append(ra)
        ranges = sorted(ranges, key=lambda x: len(x.missing))
        for ra in ranges:
            ra.add(dsp)

    def assemble(self, compact=1):
        cells, get = {}, sh.get_nested_dicts
        for c in self.cells.values():
            if isinstance(c, Ref):
                continue
            rng = c.range.ranges[0]
            indices = RangesAssembler._range_indices(c.range)
            if len(indices) == 1:
                get(cells, 'cell', rng['sheet_id'])[list(indices)[0]] = c.output
            else:
                get(cells, 'range', rng['sheet_id'], default=list).append(
                    (c.output, indices)
                )

        self._assemble_ranges(cells, compact=compact)
        return self

    def inverse_references(self):
        dsp = self.dsp
        pred, succ, nodes = dsp.dmap.pred, dsp.dmap.succ, dsp.nodes
        for c in tuple(self.cells.values()):
            if isinstance(c, Ref) and c.inputs:
                if c.func.dsp.function_nodes:
                    continue
                inp = c.output
                if set(pred[inp]) == {c.func.function_id}:
                    out = list(c.inputs)[0]
                    if not any(out in succ[k] for k in succ[inp]):
                        dsp.add_function(
                            '=%s' % inp, sh.bypass, inputs=[inp], outputs=[out]
                        )
                        d = nodes[inp]
                        d['inv-data'] = {out}
                        if 'filters' in nodes[out]:
                            sh.get_nested_dicts(
                                d, 'filters', default=list
                            ).extend(nodes[out]['filters'])

    def finish(self, complete=True, circular=False, assemble=True):
        if complete:
            self.complete()
        if assemble:
            self.assemble()
        if circular:
            self.solve_circular()
        self.inverse_references()
        return self

    def to_dict(self):
        nodes = {
            k: d['value']
            for k, d in self.dsp.default_values.items()
            if not isinstance(k, sh.Token)
        }
        nodes = {
            k: isinstance(v, str) and v.startswith('=') and '="%s"' % v or v
            for k, v in nodes.items()
        }
        nodes = {
            k: '#EMPTY' if v == [[sh.EMPTY]] else v
            for k, v in nodes.items()
        }
        nodes = {
            k: {
                'type': 'HexValue', 'value': v
            } if isinstance(v, HexValue) else v
            for k, v in nodes.items()
        }
        for d in self.dsp.function_nodes.values():
            fun = d['function']
            if isinstance(fun, CellWrapper):
                nodes.update(dict.fromkeys(d['outputs'], fun.__name__))
        return nodes

    def from_dict(self, adict, context=None, assemble=True, ref=True):
        refs, cells, nodes, get = {}, {}, set(), sh.get_nested_dicts
        for k, v in adict.items():
            if isinstance(v, dict):
                if v['type'] == 'HexValue':
                    v = HexValue(v['value'])
            if isinstance(v, str) and v.upper() == '#EMPTY':
                v = [[sh.EMPTY]]
            try:
                cell = Cell(k, v, context=context, replace_missing_ref=ref)
            except ValueError:
                cell = Ref(k, v, context=context).compile(context=context)
                refs[cell.output] = None
                nodes.update(cell.add(self.dsp, context=context))
            cells[cell.output] = cell
        self._update_refs(nodes, refs)
        for k, cell in cells.items():
            if k not in refs:
                nodes.update(cell.compile(references=refs).add(self.dsp))
        self.cells.update(cells)
        if assemble:
            self.assemble()
        self.inverse_references()
        return self

    def write(self, books=None, solution=None, dirpath=None):
        books = {} if books is None else books
        solution = self.dsp.solution if solution is None else solution
        are_in, get_in = sh.are_in_nested_dicts, sh.get_nested_dicts
        for k, r in solution.items():
            if isinstance(k, sh.Token):
                continue
            if isinstance(r, Ranges):
                rng = {k: v for k, v in _re_sheet_id.match(
                    r.ranges[0]['sheet_id']
                ).groupdict().items() if v is not None}
                rng.update(r.ranges[0])
            else:
                try:
                    r = Ranges().push(k, r)
                    rng = r.ranges[0]
                except ValueError:  # Reference.
                    rng = {'sheet': ''}
            fpath = _encode_path(osp.join(
                _decode_path(rng.get('directory', '')), rng.get('filename', '')
            ))
            fpath, sheet_name = _get_name(fpath, books), rng.get('sheet')
            if not (fpath and sheet_name):
                log.info('Node `%s` cannot be saved '
                         '(missing filename and/or sheet_name).' % k)
                continue
            elif _re_build_id.match(fpath):
                continue
            if not are_in(books, fpath, BOOK):
                from openpyxl import Workbook
                book = get_in(books, fpath, BOOK, default=Workbook)
                for ws in book.worksheets:
                    book.remove(ws)
            else:
                book = books[fpath][BOOK]

            sheet_names = book.sheetnames
            sheet_name = _get_name(sheet_name, sheet_names)
            if sheet_name not in sheet_names:
                book.create_sheet(sheet_name)
            sheet = book[sheet_name]
            rng['c1'] = rng['c1'] or 'A'
            rng['r1'] = int(rng['r1']) or 1
            ref = '{c1}{r1}:{c2}{r2}'.format(**rng)
            for c, v in zip(np.ravel(sheet[ref]), np.ravel(r.value)):
                try:
                    if v is sh.EMPTY:
                        v = None
                    elif isinstance(v, np.generic):
                        v = v.item()
                    elif isinstance(v, XlError):
                        v = str(v)
                    c.value = v
                    if c.data_type == 'f':
                        c.data_type = 's'
                except AttributeError:
                    pass
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
            for fpath, d in books.items():
                d[BOOK].save(osp.join(dirpath, _decode_path(fpath)))
        return books

    def compile(self, inputs, outputs):
        dsp = self.dsp.shrink_dsp(inputs=inputs, outputs=outputs)
        inp = set(inputs)
        nodes = dsp.nodes
        for i in inputs:
            inp.update(nodes.get(i, {}).get('inv-data', ()))
        dsp.default_values = {
            k: v for k, v in dsp.default_values.items() if k not in inp
        }

        res = dsp()

        dsp = dsp.get_sub_dsp_from_workflow(
            outputs, graph=dsp.dmap, reverse=True, blockers=res,
            wildcard=False
        )

        for k, v in res.items():
            if k in dsp.data_nodes and k not in dsp.default_values:
                dsp.set_default_value(k, v.value)

        func = self.compile_class(
            dsp=dsp,
            function_id=self.dsp.name,
            inputs=inputs,
            outputs=outputs
        )

        return func

    def solve_circular(self):
        from .cycle import simple_cycles
        from collections import Counter
        mod, dsp = {}, self.dsp
        f_nodes, d_nodes, dmap = dsp.function_nodes, dsp.data_nodes, dsp.dmap
        skip_nodes = {
            k for k, node in f_nodes.items()
            if isinstance(node['function'], InvRangesAssembler)
        }

        cycles = list(simple_cycles(dmap.succ, skip_nodes=skip_nodes))
        cycles_nodes = Counter(sum(cycles, []))
        for cycle in sorted(map(set, cycles)):
            cycles_nodes.subtract(cycle)
            active_nodes = {k for k, v in cycles_nodes.items() if v}
            for k in sorted(cycle.intersection(f_nodes)):
                if _check_cycles(dmap, k, f_nodes, cycle, active_nodes, mod):
                    break
            else:
                cycles_nodes.update(cycle)
                dist = sh.inf(len(cycle) + 1, 0)
                for k in sorted(cycle.intersection(d_nodes)):
                    dsp.set_default_value(k, ERR_CIRCULAR, dist)

        if mod:  # Update dsp.
            dsp.add_data(CIRCULAR, ERR_CIRCULAR)

            for k, v in mod.items():
                d = f_nodes[k]
                d['inputs'] = [CIRCULAR if i in v else i for i in d['inputs']]
                dmap.remove_edges_from(((i, k) for i in v))
                dmap.add_edge(CIRCULAR, k)

        return self


def _check_range_all_cycles(nodes, active_nodes, j):
    if isinstance(nodes[j]['function'], RangesAssembler):
        return active_nodes.intersection(nodes[j]['inputs'])
    return False


def _check_cycles(dmap, node_id, nodes, cycle, active_nodes, mod=None):
    node, mod = nodes[node_id], {} if mod is None else mod
    _map = dict(zip(node['function'].inputs, node['inputs']))
    pred, res = dmap.pred, ()
    check = functools.partial(_check_range_all_cycles, nodes, active_nodes)
    if not any(any(map(check, pred[k])) for k in _map.values() if k in cycle):
        cycle = [i for i, j in _map.items() if j in cycle]
        try:
            res = tuple(map(_map.get, node['function'].check_cycles(cycle)))
            res and sh.get_nested_dicts(mod, node_id, default=set).update(res)
        except AttributeError:
            pass
    return res
